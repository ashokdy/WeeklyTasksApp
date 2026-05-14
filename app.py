#!/usr/bin/env python3
"""
MAG Weekly Tasks Web App
Multi-user task tracker with natural language input + weekly Excel reports
"""
import os, re, io, sqlite3
from datetime import datetime, date, timedelta
from flask import (Flask, render_template_string, request, redirect,
                   url_for, session, send_file, flash, jsonify)
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.fills import PatternFill as PF

app = Flask(__name__)
app.secret_key = 'mag-weekly-tasks-2026-secret'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)  # stay logged in for 30 days

# ── DB path: prefer ~/Library/Application Support/ (no TCC issues on macOS),
#    fall back to app directory, then /tmp as last resort ──────────────────────
def _resolve_db_path():
    candidates = []
    # 1. macOS Application Support — safest, no TCC restriction
    if os.uname().sysname == 'Darwin':
        app_support = os.path.expanduser('~/Library/Application Support/MAGWeeklyTasks')
        candidates.append(os.path.join(app_support, 'tasks.db'))
    # 2. Same directory as app.py
    candidates.append(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'tasks.db'))
    # 3. /tmp fallback (works everywhere, not persistent across reboots but fine for demo)
    candidates.append('/tmp/mag_weeklytasks.db')

    for path in candidates:
        folder = os.path.dirname(path)
        try:
            os.makedirs(folder, exist_ok=True)
            # Quick write test
            test = path + '.writetest'
            with open(test, 'w') as f:
                f.write('ok')
            os.remove(test)
            print(f"  DB location: {path}")
            return path
        except Exception:
            continue
    raise RuntimeError("No writable location found for the database.")

DB = _resolve_db_path()

# ── Database ──────────────────────────────────────────────────────────────────

def db():
    c = sqlite3.connect(DB, check_same_thread=False)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    conn = db()
    try:
        conn.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            task_id TEXT,
            type TEXT DEFAULT 'General & Admin',
            description TEXT,
            requested_by TEXT,
            assignee TEXT,
            creation_date DATE,
            resolved_on DATE,
            priority TEXT DEFAULT 'Low',
            status TEXT DEFAULT 'Completed',
            hours_spent REAL DEFAULT 0,
            notes TEXT,
            raw_input TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )''')
        conn.commit()
        # Auto-create default account if no users exist (survives DB resets)
        existing = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
        if existing == 0:
            from werkzeug.security import generate_password_hash as _gph
            conn.execute(
                'INSERT INTO users (username, password_hash, full_name) VALUES (?,?,?)',
                ('ashokdy', _gph('Mag2026', method='pbkdf2:sha256'), 'Ashok Devangam Yerra')
            )
            conn.commit()
            print("  Default user created: ashokdy / Mag2026")
        # Migration: fix full_name for ashokdy if it was stored as a short/incorrect value
        row = conn.execute("SELECT id, full_name FROM users WHERE username='ashokdy'").fetchone()
        if row and row['full_name'] != 'Ashok Devangam Yerra':
            old_name = row['full_name']
            conn.execute("UPDATE users SET full_name='Ashok Devangam Yerra' WHERE username='ashokdy'")
            conn.execute(
                "UPDATE tasks SET requested_by='Ashok Devangam Yerra' WHERE user_id=? AND requested_by=?",
                (row['id'], old_name)
            )
            conn.execute(
                "UPDATE tasks SET assignee='Ashok Devangam Yerra' WHERE user_id=? AND assignee=?",
                (row['id'], old_name)
            )
            conn.commit()
            print(f"  Migrated full_name from '{old_name}' to 'Ashok Devangam Yerra'")
        print(f"  Database ready: {DB}")
    finally:
        conn.close()

# Always init DB when module loads (covers both direct run and service launch)
init_db()

# ── Natural Language Parser ───────────────────────────────────────────────────

MONTHS = {
    'jan':1,'january':1,'feb':2,'february':2,'mar':3,'march':3,
    'apr':4,'april':4,'may':5,'jun':6,'june':6,'jul':7,'july':7,
    'aug':8,'august':8,'sep':9,'september':9,'oct':10,'october':10,
    'nov':11,'november':11,'dec':12,'december':12
}

def group_task_lines(raw):
    """
    Handles both input formats:
      Multi-line: Task ID alone on one line, description on the next line(s)
      Single-line: Everything on one line per task
    Returns a list of single-line strings, one per task, ready for parse_input().
    """
    tid_start = re.compile(r'^([A-Za-z]{2,6}\d+|Adhoc)\b', re.I)
    lines = [l.strip() for l in raw.strip().splitlines() if l.strip()]
    groups, current = [], []
    for line in lines:
        if tid_start.match(line):
            if current:
                groups.append(' '.join(current))
            current = [line]
        else:
            current.append(line)
    if current:
        groups.append(' '.join(current))
    return groups if groups else lines

def parse_date_str(s):
    s = s.strip().lower()
    yr = datetime.now().year
    m = re.match(r'^(\d{1,2})\s+(\w+)(?:\s+(\d{4}))?$', s)
    if m:
        day, mon, y = m.groups()
        if mon in MONTHS:
            return date(int(y or yr), MONTHS[mon], int(day))
    m = re.match(r'^(\w+)\s+(\d{1,2})(?:\s+(\d{4}))?$', s)
    if m:
        mon, day, y = m.groups()
        if mon in MONTHS:
            return date(int(y or yr), MONTHS[mon], int(day))
    return None

def extract_date_hours(text):
    """Return list of (date, hours) pairs and cleaned description text."""
    tl = text.lower()
    mon_re = '|'.join(sorted(MONTHS, key=len, reverse=True))
    date_re = rf'(?:(\d{{1,2}})\s+(?:{mon_re})|(?:{mon_re})\s+(\d{{1,2}}))'
    pairs, spans = [], []

    # Pattern A: "on DATE for N [hours]"
    pA = rf'on\s+({date_re})\s+for\s+(\d+(?:\.\d+)?)\s*(?:hours?)?'
    for m in re.finditer(pA, tl):
        d = parse_date_str(m.group(1))
        h = float(m.group(m.lastindex))
        if d:
            pairs.append((d, h, m.span()))
            spans.append(m.span())

    # Pattern B: "N [hours] on DATE"
    pB = rf'(\d+(?:\.\d+)?)\s*(?:hours?)?\s+on\s+({date_re})'
    for m in re.finditer(pB, tl):
        if any(s[0] <= m.start() <= s[1] for s in spans):
            continue
        d = parse_date_str(m.group(2))
        h = float(m.group(1))
        if d:
            pairs.append((d, h, m.span()))
            spans.append(m.span())

    # Pattern C: "DATE for N [hours]"  (no "on" prefix)
    pC = rf'({date_re})\s+for\s+(\d+(?:\.\d+)?)\s*(?:hours?)?'
    for m in re.finditer(pC, tl):
        if any(s[0] <= m.start() <= s[1] for s in spans):
            continue
        d = parse_date_str(m.group(1))
        h = float(m.group(m.lastindex))
        if d:
            pairs.append((d, h, m.span()))
            spans.append(m.span())

    pairs.sort(key=lambda x: x[0])

    # Strip matched spans from text
    clean = text
    for span in sorted([p[2] for p in pairs], reverse=True):
        clean = clean[:span[0]] + ' ' + clean[span[1]:]

    return [(d, h) for d, h, _ in pairs], re.sub(r'\s+', ' ', clean).strip()

def parse_input(raw, full_name, default_date=None):
    """Parse natural language into list of task dicts (one per date)."""
    text = raw.strip()

    # Task ID
    tid_m = re.match(r'^([A-Za-z]{1,4}\d+|Adhoc|adhoc|ADHOC)\s*', text, re.I)
    task_id = None
    if tid_m:
        task_id = 'Adhoc' if tid_m.group(1).lower() == 'adhoc' else tid_m.group(1).upper()
        text = text[tid_m.end():]

    # Type
    type_val = 'General & Admin'
    tm = re.search(r'\btype\s*[-:]\s*([A-Za-z&/ ]+?)(?=\s{2,}|\s+\w+\s+\w|\s*$)', text, re.I)
    if tm:
        type_val = tm.group(1).strip()
        text = text[:tm.start()] + text[tm.end():]
    elif re.search(r'\bR&D\b|\bRnD\b', text, re.I):
        type_val = 'R&D'
        text = re.sub(r'\bR&D\b|\bRnD\b', '', text, flags=re.I)

    # Status
    status = 'Completed'
    sm = re.search(r'\b(in\s+progress|completed?|pending|on\s+hold)\b', text, re.I)
    if sm:
        sl = sm.group(1).lower()
        if 'progress' in sl: status = 'In Progress'
        elif 'complet' in sl: status = 'Completed'
        elif 'pending' in sl: status = 'Pending'
        elif 'hold' in sl: status = 'On Hold'
        text = text[:sm.start()] + text[sm.end():]

    # Priority
    priority = 'Low'
    pm = re.search(r'\b(low|medium|high|critical)\s*priority\b', text, re.I)
    if pm:
        priority = pm.group(1).capitalize()
        text = text[:pm.start()] + text[pm.end():]

    # Notes / comments
    notes = None
    nm = re.search(r'(?:with\s+)?(?:comments?|notes?)\s+(?:as\s+)?["\']([^"\']+)["\']', text, re.I)
    if not nm:
        nm = re.search(r'(?:comments?|notes?)\s*[:\-]\s*(.+?)(?=\s+(?:on|and)\s+\d|\s*$)', text, re.I)
    if nm:
        notes = nm.group(1).strip().strip('"\'')
        text = text[:nm.start()] + text[nm.end():]

    # Date-hour pairs
    date_hours, remaining = extract_date_hours(text)

    # Clean description — strip any number of trailing/leading connector words
    desc = re.sub(r'(\s*\b(and|with|for|on)\b)+\s*$', '', remaining, flags=re.I)
    desc = re.sub(r'^\s*(\b(and|with|for|on)\b\s*)+', '', desc, flags=re.I)
    desc = re.sub(r'\s+', ' ', desc).strip(' ,-')

    if not date_hours:
        date_hours = [(default_date or date.today(), 0)]

    # Build one entry per date, sorted by creation_date
    entries = []
    for i, (cdate, hours) in enumerate(date_hours):
        resolved = cdate + timedelta(days=1)
        entries.append({
            'task_id':      task_id or 'Adhoc',
            'type':         type_val,
            'description':  desc,
            'requested_by': full_name,
            'assignee':     full_name,
            'creation_date': cdate.isoformat(),
            'resolved_on':  resolved.isoformat(),
            'priority':     priority,
            'status':       status,
            'hours_spent':  hours,
            'notes':        notes if i == len(date_hours) - 1 else None,
            'raw_input':    raw,
        })
    return entries

# ── Week helpers ──────────────────────────────────────────────────────────────

def week_bounds(for_date=None):
    """Return (monday, friday) of the week containing for_date."""
    d = for_date or date.today()
    mon = d - timedelta(days=d.weekday())
    fri = mon + timedelta(days=4)
    return mon, fri

def friday_of(d=None):
    d = d or date.today()
    return d + timedelta(days=(4 - d.weekday()) % 7)

# ── Excel report ──────────────────────────────────────────────────────────────

COLS = ['Task ID','Type','Description','Requested By','Assignee',
        'Creation Date','Resolved On','Priority','Status','Hours Spent','CR Number / Notes']
COL_W = [10, 16, 55, 22, 22, 14, 14, 10, 15, 13, 45]

def make_report(week_fri):
    """Generate Excel workbook with one sheet per user for the given week."""
    mon, fri = week_bounds(week_fri)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font  = Font(bold=True, size=11, name='Calibri')
    hdr_fill  = PF(fill_type='solid', fgColor='2E75B6')
    hdr_font2 = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='BFBFBF')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PF(fill_type='solid', fgColor='EBF3FB')

    conn = db()
    users = conn.execute('SELECT * FROM users ORDER BY full_name').fetchall()
    for u in users:
        rows = conn.execute('''
            SELECT * FROM tasks
            WHERE user_id=? AND creation_date BETWEEN ? AND ?
            ORDER BY creation_date ASC, id ASC
        ''', (u['id'], mon.isoformat(), fri.isoformat())).fetchall()

        if not rows:
            continue

        sheet_name = u['full_name'][:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header
        ws.append(COLS)
        for ci, cell in enumerate(ws[1], 1):
            cell.font   = hdr_font2
            cell.fill   = hdr_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = COL_W[ci-1]
        ws.row_dimensions[1].height = 22

        # Data rows
        total_hours = 0
        for ri, row in enumerate(rows, 2):
            ws.append([
                row['task_id'], row['type'], row['description'],
                row['requested_by'], row['assignee'],
                row['creation_date'], row['resolved_on'],
                row['priority'], row['status'],
                row['hours_spent'], row['notes']
            ])
            for ci, cell in enumerate(ws[ri], 1):
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=(ci==3))
                if ri % 2 == 0:
                    cell.fill = alt_fill
                if ci in (6, 7) and cell.value:
                    cell.number_format = 'DD/MM/YYYY'
            ws.row_dimensions[ri].height = 18
            total_hours += row['hours_spent'] or 0

        # Totals row
        tr = len(rows) + 2
        ws.cell(tr, 9, 'TOTAL HOURS').font  = Font(bold=True)
        ws.cell(tr, 9).fill   = PF(fill_type='solid', fgColor='D9E1F2')
        ws.cell(tr, 10, total_hours).font = Font(bold=True)
        ws.cell(tr, 10).fill  = PF(fill_type='solid', fgColor='D9E1F2')
        ws.freeze_panes = 'A2'

    conn.close()
    if not wb.sheetnames:
        ws = wb.create_sheet('No Data')
        ws['A1'] = 'No tasks recorded for this week.'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── HTML Templates ────────────────────────────────────────────────────────────

BASE = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>MAG Weekly Tasks</title>
<link href="https://cdnjs.cloudflare.com/ajax/libs/bootstrap/5.3.2/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdnjs.cloudflare.com/ajax/libs/bootstrap-icons/1.11.3/font/bootstrap-icons.min.css" rel="stylesheet">
<style>
  :root{--mag:#1a4fa0;--mag-lt:#2E75B6;--accent:#e8f0fe;}
  body{background:#f4f6fb;font-family:'Segoe UI',sans-serif;}
  .navbar{background:var(--mag)!important;}
  .navbar-brand,.nav-link,.navbar-text{color:#fff!important;}
  .card{border:none;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.07);}
  .card-header{background:var(--mag);color:#fff;border-radius:12px 12px 0 0!important;font-weight:600;}
  .btn-primary{background:var(--mag);border-color:var(--mag);}
  .btn-primary:hover{background:var(--mag-lt);border-color:var(--mag-lt);}
  .badge-status-completed{background:#198754;}
  .badge-status-in-progress{background:#fd7e14;}
  .badge-status-pending{background:#6c757d;}
  .task-input{font-family:monospace;font-size:.95rem;}
  .week-pill{background:var(--accent);color:var(--mag);border-radius:20px;
             padding:4px 14px;font-size:.85rem;font-weight:600;}
  .hours-badge{background:var(--mag);color:#fff;border-radius:8px;
               padding:2px 10px;font-weight:700;}
  .tbl thead th{background:var(--mag);color:#fff;font-weight:600;font-size:.85rem;}
  .tbl tbody tr:hover{background:#f0f5ff;}
  .hint{font-size:.8rem;color:#6c757d;}
  .flash-success{background:#d1e7dd;color:#0a3622;border-radius:8px;padding:10px 16px;}
  .flash-error{background:#f8d7da;color:#58151c;border-radius:8px;padding:10px 16px;}
  .entry-card{border-left:4px solid var(--mag-lt);background:#fff;border-radius:0 8px 8px 0;
              padding:10px 14px;margin-bottom:8px;}
  .ie-field{border:1px solid transparent;background:transparent;box-shadow:none;
            padding:2px 4px;border-radius:4px;transition:border-color .15s,background .15s;}
  .ie-field:hover{border-color:#dee2e6!important;}
  .status-sel{cursor:pointer;font-size:.75rem;padding:2px 10px;border-radius:12px;
              font-weight:600;border:none;outline:none;transition:background .2s;}
  .s-Completed{background:#d1e7dd;color:#0a3622;}
  .s-In-Progress{background:#fff3cd;color:#664d03;}
  .s-Pending{background:#e2e3e5;color:#383d41;}
  .s-On-Hold{background:#f8d7da;color:#58151c;}
</style>
</head>
<body>
<nav class="navbar navbar-expand-lg px-3 py-2">
  <a class="navbar-brand fw-bold fs-5" href="/"><i class="bi bi-clipboard-data me-2"></i>MAG Weekly Tasks</a>
  <div class="ms-auto d-flex align-items-center gap-3">
    {% if session.user_id %}
    <span class="navbar-text opacity-75">Hi, {{ session.full_name.split()[0] }}</span>
    <a href="/logout" class="btn btn-sm btn-outline-light">Logout</a>
    {% endif %}
  </div>
</nav>
<div class="container py-4">
{% with msgs = get_flashed_messages(with_categories=true) %}
  {% for cat,msg in msgs %}
  <div class="flash-{{cat}} mb-3">{{ msg }}</div>
  {% endfor %}
{% endwith %}
{% block content %}{% endblock %}
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/bootstrap/5.3.2/js/bootstrap.bundle.min.js"></script>
{% block scripts %}{% endblock %}
</body>
</html>'''

LOGIN_T = BASE.replace('{% block content %}{% endblock %}', '''
<div class="row justify-content-center mt-5">
<div class="col-md-5">
  <div class="card">
    <div class="card-header text-center py-3"><i class="bi bi-person-lock me-2"></i>Sign In</div>
    <div class="card-body p-4">
      <form method="post">
        <div class="mb-3">
          <label class="form-label fw-semibold">Username</label>
          <input name="username" class="form-control" placeholder="your.username" required autofocus>
        </div>
        <div class="mb-4">
          <label class="form-label fw-semibold">Password</label>
          <input name="password" type="password" class="form-control" required>
        </div>
        <button class="btn btn-primary w-100 py-2">Sign In</button>
      </form>
      <hr>
      <p class="text-center hint mb-1">No account? <a href="/register">Register here</a></p>
      <div class="text-center hint" style="font-size:.75rem;color:#999;">
        Default: <code>ashokdy</code> / <code>Mag2026</code>
      </div>
    </div>
  </div>
</div></div>
''').replace('{% block scripts %}{% endblock %}', '')

REGISTER_T = BASE.replace('{% block content %}{% endblock %}', '''
<div class="row justify-content-center mt-5">
<div class="col-md-5">
  <div class="card">
    <div class="card-header text-center py-3"><i class="bi bi-person-plus me-2"></i>Create Account</div>
    <div class="card-body p-4">
      <form method="post">
        <div class="mb-3">
          <label class="form-label fw-semibold">Full Name</label>
          <input name="full_name" class="form-control" placeholder="Ashok Devangam Yerra" required autofocus>
        </div>
        <div class="mb-3">
          <label class="form-label fw-semibold">Username</label>
          <input name="username" class="form-control" placeholder="a.yerra" required>
        </div>
        <div class="mb-4">
          <label class="form-label fw-semibold">Password</label>
          <input name="password" type="password" class="form-control" required>
        </div>
        <button class="btn btn-primary w-100 py-2">Create Account</button>
      </form>
      <hr><p class="text-center hint mb-0">Already have an account? <a href="/login">Sign in</a></p>
    </div>
  </div>
</div></div>
''').replace('{% block scripts %}{% endblock %}', '')

DASHBOARD_T = BASE.replace('{% block content %}{% endblock %}', '''
<!-- Week header -->
<div class="d-flex align-items-center justify-content-between mb-4 flex-wrap gap-2">
  <div>
    <h5 class="mb-0 fw-bold text-dark">Dashboard</h5>
    <span class="week-pill mt-1 d-inline-block">
      Week: {{ mon.strftime("%d %b") }} – {{ fri.strftime("%d %b %Y") }}
    </span>
  </div>
  <div class="d-flex gap-2 align-items-center flex-wrap">
    <form method="get" class="d-flex gap-2 align-items-center">
      <label class="hint mb-0">View week:</label>
      <select name="week" class="form-select form-select-sm" style="width:auto" onchange="this.form.submit()">
        {% for w in weeks %}
        <option value="{{ w }}" {% if w==sel_week %}selected{% endif %}>
          {{ w }}{% if w == fri.isoformat() %} (current){% endif %}
        </option>
        {% endfor %}
      </select>
    </form>
    <a href="/report?week={{ fri.isoformat() }}" class="btn btn-success btn-sm">
      <i class="bi bi-file-earmark-excel me-1"></i>Weekly Report
    </a>
    <a href="/clear-week/{{ fri.isoformat() }}"
       class="btn btn-outline-danger btn-sm"
       onclick="return confirm('Delete ALL your entries for week {{ mon.strftime(\'%d %b\') }}–{{ fri.strftime(\'%d %b\') }}?')">
      <i class="bi bi-trash me-1"></i>Clear Week
    </a>
  </div>
</div>

<!-- Input card -->
<div class="card mb-4">
  <div class="card-header"><i class="bi bi-plus-circle me-2"></i>Add Tasks (Natural Language)</div>
  <div class="card-body p-4">
    <form method="post" action="/add">
      <input type="hidden" name="default_date" value="{{ mon.isoformat() }}">
      <textarea name="raw" class="form-control task-input mb-2" rows="3"
        placeholder="TSK24645 Super+: Leasing enhancements in progress 3 on 8 Apr and 3 on 9 Apr&#10;TSK24642 One+: Google Location Policy changes 3 on 8 Apr 5 on 9 Apr 7 on 10 Apr with comments as &quot;changes are done, testing in all scenarios&quot;&#10;Adhoc Type - R&D Studied background location usage in RN on Apr 8 for 3 hours" required></textarea>
      <div class="d-flex align-items-center gap-3">
        <button class="btn btn-primary px-4"><i class="bi bi-send me-1"></i>Add Entries</button>
        <span class="hint">One task per line · All fields are optional except description</span>
      </div>
    </form>
    <div class="mt-3 p-3 rounded" style="background:#f8f9ff;font-size:.8rem;color:#555;">
      <strong>Rules applied automatically:</strong>
      Resolved Date = Created + 1 day if not provided &nbsp;·&nbsp;
      Multiple dates = separate rows &nbsp;·&nbsp;
      Entries sorted by Created Date &nbsp;·&nbsp;
      Defaults: Type=General &amp; Admin, Priority=Low, Status=Completed
    </div>
  </div>
</div>

<!-- This week's tasks -->
<div class="card">
  <div class="card-header d-flex justify-content-between align-items-center">
    <div class="d-flex align-items-center gap-2">
      <span><i class="bi bi-table me-2"></i>This Week's Tasks</span>
      <button type="button" class="btn btn-sm btn-light py-0 px-2" onclick="addNewRow()" title="Add a new row directly in the table">
        <i class="bi bi-plus-lg"></i>
      </button>
    </div>
    <span class="hours-badge" id="hoursTotal">{{ total_hours }}h total</span>
  </div>
  <div class="card-body p-0">
    <form method="post" action="/delete-selected" id="bulkForm">
    {% if tasks %}
    <div class="d-flex gap-2 px-3 pt-2 pb-1 border-bottom">
      <button type="button" class="btn btn-sm btn-outline-secondary" onclick="toggleAll()">
        <i class="bi bi-check2-square me-1"></i>Select All
      </button>
      <button type="submit" class="btn btn-sm btn-danger"
              onclick="return document.querySelectorAll('.row-check:checked').length>0 || (alert('Select at least one row.'),false)">
        <i class="bi bi-trash me-1"></i>Delete Selected
      </button>
    </div>
    {% endif %}
    <div class="table-responsive">
    <table class="table tbl table-hover mb-0 small">
      <thead>
        <tr>
          <th style="width:36px"></th>
          <th>Task ID</th><th>Type</th><th style="min-width:200px">Description</th>
          <th>Created</th><th>Resolved</th><th>Status</th><th>Hours</th><th>Notes</th>
          <th></th>
        </tr>
      </thead>
      <tbody id="tasksTbody">
      <datalist id="typeListMain">
        <option value="General &amp; Admin"><option value="R&amp;D">
        <option value="Development"><option value="Support"><option value="Testing">
      </datalist>
      {% for t in tasks %}
        <tr>
          <td><input type="checkbox" name="ids" value="{{ t.id }}" class="form-check-input row-check"></td>
          <td>
            <input type="text" class="ie-field form-control form-control-sm fw-semibold"
                   data-id="{{ t.id }}" data-field="task_id"
                   value="{{ t.task_id }}" style="min-width:72px">
          </td>
          <td>
            <input type="text" class="ie-field form-control form-control-sm"
                   data-id="{{ t.id }}" data-field="type"
                   value="{{ t.type }}" list="typeListMain" style="min-width:100px">
          </td>
          <td>
            <input type="text" class="ie-field form-control form-control-sm"
                   data-id="{{ t.id }}" data-field="description"
                   value="{{ t.description }}" style="min-width:190px">
          </td>
          <td>
            <input type="date" class="ie-field form-control form-control-sm"
                   data-id="{{ t.id }}" data-field="creation_date"
                   value="{{ t.creation_date }}" style="min-width:118px">
          </td>
          <td>
            <input type="date" class="ie-field form-control form-control-sm"
                   data-id="{{ t.id }}" data-field="resolved_on"
                   value="{{ t.resolved_on }}" style="min-width:118px">
          </td>
          <td>
            <select class="status-sel" data-id="{{ t.id }}" data-field="status">
              <option {% if t.status=='Completed' %}selected{% endif %}>Completed</option>
              <option {% if t.status=='In Progress' %}selected{% endif %}>In Progress</option>
              <option {% if t.status=='Pending' %}selected{% endif %}>Pending</option>
              <option {% if t.status=='On Hold' %}selected{% endif %}>On Hold</option>
            </select>
          </td>
          <td class="text-center">
            <input type="number" step="0.5" min="0" max="24"
                   class="hours-edit form-control form-control-sm text-center fw-bold p-0"
                   value="{{ t.hours_spent }}" data-id="{{ t.id }}"
                   style="width:68px;margin:auto;border:1px solid transparent;background:transparent;box-shadow:none;">
          </td>
          <td>
            <input type="text" class="ie-field form-control form-control-sm"
                   data-id="{{ t.id }}" data-field="notes"
                   value="{{ t.notes or '' }}" placeholder="—"
                   style="min-width:110px;color:#555;font-size:.8rem;">
          </td>
          <td class="text-nowrap">
            <a href="/clone/{{ t.id }}?week={{ fri.isoformat() }}" class="btn btn-outline-secondary btn-sm py-0 me-1"
               title="Clone this row">
              <i class="bi bi-copy"></i>
            </a>
            <a href="/delete/{{ t.id }}?week={{ fri.isoformat() }}" class="btn btn-outline-danger btn-sm py-0"
               onclick="return confirm('Delete this entry?')">
              <i class="bi bi-trash"></i>
            </a>
          </td>
        </tr>
      {% endfor %}
      {% if not tasks %}
      <tr id="emptyRow">
        <td colspan="10" class="text-center py-4 text-muted">
          <i class="bi bi-inbox fs-2 d-block mb-1"></i>No tasks yet — click <strong>+</strong> in the header to add one.
        </td>
      </tr>
      {% endif %}
      </tbody>
      <tfoot>
        <tr class="table-info fw-bold">
          <td colspan="7" class="text-end">Total Hours</td>
          <td class="text-center" id="hoursTotalRow">{{ total_hours }}</td>
          <td colspan="2"></td>
        </tr>
      </tfoot>
    </table>
    </div>
    </form>
  </div>
</div>
''').replace('{% block scripts %}{% endblock %}', '''
<script>
// ── Textarea: Shift+Enter = newline, Enter = submit ───────────────────────────
document.querySelector('textarea').addEventListener('keydown', function(e){
  if(e.key==='Enter' && !e.shiftKey){ e.preventDefault(); this.form.submit(); }
});

// ── Bulk-select checkboxes ────────────────────────────────────────────────────
function toggleAll() {
  const boxes = document.querySelectorAll('.row-check');
  const allChecked = [...boxes].every(b => b.checked);
  boxes.forEach(b => b.checked = !allChecked);
}

// ── Hours total recalc ────────────────────────────────────────────────────────
function recalcTotal() {
  let total = 0;
  document.querySelectorAll('.hours-edit').forEach(i => total += parseFloat(i.value) || 0);
  const t = Math.round(total * 10) / 10;
  const badge = document.getElementById('hoursTotal');
  const row   = document.getElementById('hoursTotalRow');
  if (badge) badge.textContent = t + 'h total';
  if (row)   row.textContent   = t;
}

// ── Shared helpers ────────────────────────────────────────────────────────────
function flashCell(el, ok) {
  el.style.border     = ok ? '1px solid #198754' : '1px solid #dc3545';
  el.style.background = ok ? '#f0fff4' : '#fff0f0';
  setTimeout(() => { el.style.borderColor = 'transparent'; el.style.background = 'transparent'; },
             ok ? 900 : 1400);
}
function saveField(el, field, value, orig, onSuccess) {
  fetch('/update/' + el.dataset.id, {
    method: 'POST', headers: {'Content-Type': 'application/x-www-form-urlencoded'},
    body: 'field=' + encodeURIComponent(field) + '&value=' + encodeURIComponent(value)
  }).then(r => r.json()).then(d => {
    if (d.ok) { flashCell(el, true); if (onSuccess) onSuccess(); }
    else      { el.value = orig; flashCell(el, false); }
  }).catch(() => { el.value = orig; });
}

// ── Hours (number input, keeps totals live) ───────────────────────────────────
document.querySelectorAll('.hours-edit').forEach(input => {
  const orig = input.value;
  input.addEventListener('focus', function() {
    this.style.border = '1px solid #2E75B6'; this.style.background = '#fff';
  });
  input.addEventListener('blur', function() {
    const val = parseFloat(this.value);
    if (isNaN(val) || val < 0 || val > 24) { this.value = orig; flashCell(this, false); return; }
    saveField(this, 'hours_spent', val, orig, () => { this.value = val; recalcTotal(); });
  });
  input.addEventListener('keydown', function(e) {
    if (e.key==='Enter')  { e.preventDefault(); this.blur(); }
    if (e.key==='Escape') { this.value = orig; this.blur(); }
  });
});

// ── Universal inline field editor (text, date, notes, task_id, etc.) ─────────
document.querySelectorAll('.ie-field').forEach(input => {
  const orig  = input.value;
  const field = input.dataset.field;

  input.addEventListener('focus', function() {
    this.style.borderColor = '#2E75B6'; this.style.background = '#fff';
  });
  input.addEventListener('blur', function() {
    const val = this.value.trim();
    if (val === orig.trim()) {
      this.style.borderColor = 'transparent'; this.style.background = 'transparent'; return;
    }
    saveField(this, field, val, orig, () => {
      // When creation_date changes → auto-save resolved_on = created + 1 day
      if (field === 'creation_date') {
        const row = this.closest('tr');
        const resEl = row ? row.querySelector('[data-field="resolved_on"]') : null;
        if (resEl) {
          const d = new Date(val + 'T12:00:00');
          d.setDate(d.getDate() + 1);
          const newRes = d.toISOString().split('T')[0];
          resEl.value = newRes;
          saveField(resEl, 'resolved_on', newRes, resEl.value, null);
        }
      }
    });
  });
  input.addEventListener('keydown', function(e) {
    if (e.key==='Enter')  { e.preventDefault(); this.blur(); }
    if (e.key==='Escape') { this.value = orig; this.style.borderColor='transparent'; this.style.background='transparent'; this.blur(); }
  });
});

// ── Status select (pill style, saves on change) ───────────────────────────────
function applyStatusClass(sel) {
  sel.classList.remove('s-Completed','s-In-Progress','s-Pending','s-On-Hold');
  sel.classList.add('s-' + sel.value.replace(/\s+/g,'-'));
}
document.querySelectorAll('.status-sel').forEach(sel => {
  applyStatusClass(sel);
  sel.addEventListener('change', function() {
    const orig = [...this.options].find(o => o.defaultSelected)?.value || 'Completed';
    fetch('/update/' + this.dataset.id, {
      method: 'POST', headers: {'Content-Type': 'application/x-www-form-urlencoded'},
      body: 'field=status&value=' + encodeURIComponent(this.value)
    }).then(r => r.json()).then(d => {
      if (d.ok) { applyStatusClass(this); flashCell(this, true); }
      else      { this.value = orig; applyStatusClass(this); flashCell(this, false); }
    });
  });
});

// ── Add new inline row ────────────────────────────────────────────────────────
function addNewRow() {
  if (document.getElementById('newRow')) {
    document.getElementById('nr_task_id').focus(); return;
  }
  // Remove empty-state row if present
  const empty = document.getElementById('emptyRow');
  if (empty) empty.remove();

  // Default to Monday of the currently viewed week (so adding a row stays in that week)
  const defDate  = document.querySelector('input[name="default_date"]').value;
  const today    = defDate || new Date().toISOString().split('T')[0];
  const tomorrow = (d => { d.setDate(d.getDate()+1); return d.toISOString().split('T')[0]; })(new Date(today + 'T12:00:00'));

  const tr = document.createElement('tr');
  tr.id = 'newRow';
  tr.style.cssText = 'background:#eef4ff;';
  tr.innerHTML = `
    <td></td>
    <td><input id="nr_task_id" type="text" class="form-control form-control-sm"
               placeholder="TSK12345" style="min-width:80px"></td>
    <td><input id="nr_type" type="text" class="form-control form-control-sm"
               value="General &amp; Admin" list="typeList" style="min-width:110px">
        <datalist id="typeList">
          <option value="General &amp; Admin"><option value="R&amp;D">
          <option value="Development"><option value="Support"><option value="Testing">
        </datalist></td>
    <td><input id="nr_desc" type="text" class="form-control form-control-sm"
               placeholder="Description…" style="min-width:180px"></td>
    <td><input id="nr_created" type="date" class="form-control form-control-sm"
               value="${today}" style="min-width:120px"></td>
    <td><input id="nr_resolved" type="date" class="form-control form-control-sm"
               value="${tomorrow}" style="min-width:120px"></td>
    <td><select id="nr_status" class="form-select form-select-sm" style="min-width:110px">
          <option>Completed</option><option>In Progress</option>
          <option>Pending</option><option>On Hold</option>
        </select></td>
    <td><input id="nr_hours" type="number" step="0.5" min="0" max="24"
               class="form-control form-control-sm text-center" value="0"
               style="width:68px;margin:auto"></td>
    <td><input id="nr_notes" type="text" class="form-control form-control-sm"
               placeholder="Notes…"></td>
    <td>
      <button class="btn btn-success btn-sm py-0 me-1" type="button"
              onclick="saveNewRow()" title="Save row"><i class="bi bi-check-lg"></i></button>
      <button class="btn btn-outline-secondary btn-sm py-0" type="button"
              onclick="cancelNewRow()" title="Cancel"><i class="bi bi-x-lg"></i></button>
    </td>`;

  document.getElementById('tasksTbody').prepend(tr);

  // Auto-advance resolved date when created changes
  document.getElementById('nr_created').addEventListener('change', function() {
    const d = new Date(this.value + 'T12:00:00');
    d.setDate(d.getDate() + 1);
    document.getElementById('nr_resolved').value = d.toISOString().split('T')[0];
  });

  document.getElementById('nr_task_id').focus();
}

function cancelNewRow() {
  const r = document.getElementById('newRow');
  if (r) r.remove();
}

function saveNewRow() {
  const taskId   = document.getElementById('nr_task_id').value.trim()  || 'Adhoc';
  const type     = document.getElementById('nr_type').value.trim()      || 'General & Admin';
  const desc     = document.getElementById('nr_desc').value.trim();
  const created  = document.getElementById('nr_created').value;
  const resolved = document.getElementById('nr_resolved').value;
  const status   = document.getElementById('nr_status').value;
  const hours    = document.getElementById('nr_hours').value  || '0';
  const notes    = document.getElementById('nr_notes').value.trim();

  if (!created) { alert('Please set a Created date.'); return; }

  const saveBtn = document.querySelector('#newRow .btn-success');
  saveBtn.disabled = true; saveBtn.innerHTML = '<i class="bi bi-hourglass-split"></i>';

  fetch('/add-row', {
    method: 'POST', headers: {'Content-Type': 'application/x-www-form-urlencoded'},
    body: new URLSearchParams({task_id:taskId, type, description:desc,
          creation_date:created, resolved_on:resolved, status, hours_spent:hours, notes}).toString()
  }).then(r => r.json()).then(d => {
    if (d.ok) { window.location.href = '/?week=' + d.week; }
    else      { alert('Save failed: ' + (d.error||'unknown error'));
                saveBtn.disabled=false; saveBtn.innerHTML='<i class="bi bi-check-lg"></i>'; }
  }).catch(err => { alert('Network error: ' + err);
                    saveBtn.disabled=false; saveBtn.innerHTML='<i class="bi bi-check-lg"></i>'; });
}
</script>
''')

ALL_TASKS_T = BASE.replace('{% block content %}{% endblock %}', '''
<div class="d-flex align-items-center justify-content-between mb-4">
  <div>
    <h5 class="mb-0 fw-bold text-dark">All My Tasks</h5>
    <span class="hint">Every entry you have ever logged</span>
  </div>
  <div class="d-flex gap-2">
    <a href="/" class="btn btn-outline-primary btn-sm"><i class="bi bi-house me-1"></i>Dashboard</a>
    <form method="get" class="d-flex gap-2">
      <select name="week" class="form-select form-select-sm" onchange="this.form.submit()">
        {% for w in weeks %}
        <option value="{{ w }}" {% if w==sel_week %}selected{% endif %}>Week of {{ w }}</option>
        {% endfor %}
      </select>
      <a href="/report?week={{ sel_week }}" class="btn btn-success btn-sm">
        <i class="bi bi-file-earmark-excel me-1"></i>Report
      </a>
    </form>
  </div>
</div>
<div class="card">
  <div class="card-header d-flex justify-content-between">
    <span><i class="bi bi-table me-2"></i>Tasks — week ending {{ sel_week }}</span>
    <span class="hours-badge">{{ total_hours }}h</span>
  </div>
  <div class="card-body p-0">
    {% if tasks %}
    <div class="table-responsive">
    <table class="table tbl table-hover mb-0 small">
      <thead>
        <tr>
          <th>Task ID</th><th>Type</th><th style="min-width:220px">Description</th>
          <th>Created</th><th>Resolved</th><th>Priority</th><th>Status</th>
          <th>Hours</th><th>Notes</th><th></th>
        </tr>
      </thead>
      <tbody>
      {% for t in tasks %}
        <tr>
          <td class="fw-semibold">{{ t.task_id }}</td>
          <td><span class="badge bg-secondary">{{ t.type }}</span></td>
          <td>{{ t.description }}</td>
          <td>{{ t.creation_date }}</td>
          <td>{{ t.resolved_on }}</td>
          <td>{{ t.priority }}</td>
          <td><span class="badge badge-status-{{ t.status.lower().replace(' ','-') }}">{{ t.status }}</span></td>
          <td class="text-center fw-bold">{{ t.hours_spent }}</td>
          <td style="max-width:160px;font-size:.75rem">{{ t.notes or '' }}</td>
          <td>
            <a href="/delete/{{ t.id }}" class="btn btn-outline-danger btn-sm py-0"
               onclick="return confirm('Delete?')"><i class="bi bi-trash"></i></a>
          </td>
        </tr>
      {% endfor %}
      </tbody>
      <tfoot>
        <tr class="table-info fw-bold">
          <td colspan="7" class="text-end">Total</td>
          <td class="text-center">{{ total_hours }}</td>
          <td colspan="2"></td>
        </tr>
      </tfoot>
    </table>
    </div>
    {% else %}
    <div class="text-center py-5 text-muted">
      <i class="bi bi-inbox fs-1 d-block mb-2"></i>No tasks for this week.
    </div>
    {% endif %}
  </div>
</div>
''').replace('{% block scripts %}{% endblock %}', '')

# ── Routes ────────────────────────────────────────────────────────────────────

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect('/login')
        # Always refresh full_name from DB so stale session cookies pick up name changes
        conn = db()
        row = conn.execute('SELECT full_name FROM users WHERE id=?', (session['user_id'],)).fetchone()
        conn.close()
        if row:
            session['full_name'] = row['full_name']
        return f(*args, **kwargs)
    return decorated

@app.route('/')
@login_required
def dashboard():
    # Build list of available weeks (any week that has tasks + current week)
    conn = db()
    rows = conn.execute('''
        SELECT DISTINCT creation_date FROM tasks WHERE user_id=?
        ORDER BY creation_date DESC
    ''', (session['user_id'],)).fetchall()
    conn.close()

    seen, weeks = set(), []
    for r in rows:
        d = date.fromisoformat(r['creation_date'])
        fri = friday_of(d).isoformat()
        if fri not in seen:
            seen.add(fri)
            weeks.append(fri)
    # Always include current week and last week
    cur_fri  = friday_of().isoformat()
    last_fri = friday_of(date.today() - timedelta(days=7)).isoformat()
    if cur_fri not in seen:
        weeks.insert(0, cur_fri)
        seen.add(cur_fri)
    if last_fri not in seen:
        # Insert right after current week
        idx = weeks.index(cur_fri) + 1
        weeks.insert(idx, last_fri)
        seen.add(last_fri)

    # Selected week — default to the week with the most recent task data
    sel_week = request.args.get('week', weeks[0])
    sel_fri = date.fromisoformat(sel_week)
    sel_mon, _ = week_bounds(sel_fri)

    conn = db()
    tasks = conn.execute('''
        SELECT * FROM tasks WHERE user_id=?
        AND creation_date BETWEEN ? AND ?
        ORDER BY creation_date ASC, id ASC
    ''', (session['user_id'], sel_mon.isoformat(), sel_fri.isoformat())).fetchall()
    conn.close()
    total = sum(t['hours_spent'] or 0 for t in tasks)
    return render_template_string(DASHBOARD_T,
        tasks=tasks, mon=sel_mon, fri=sel_fri,
        total_hours=round(total, 1), weeks=weeks, sel_week=sel_week)

@app.route('/add', methods=['POST'])
@login_required
def add_task():
    raw = request.form.get('raw', '').strip()
    if not raw:
        flash('Please enter task details.', 'error')
        return redirect('/')

    # Use the week the user is currently viewing as the default date for undated entries
    try:
        default_date = date.fromisoformat(request.form.get('default_date', ''))
    except ValueError:
        default_date = date.today()

    added = 0
    skipped = 0
    errors = []
    all_dates = []  # collect all dates parsed from input to redirect correctly

    for line in group_task_lines(raw):
        line = line.strip()
        if not line:
            continue
        try:
            entries = parse_input(line, session['full_name'], default_date=default_date)
            conn = db()
            for e in entries:
                all_dates.append(date.fromisoformat(e['creation_date']))
                conn.execute('''
                    INSERT INTO tasks
                    (user_id,task_id,type,description,requested_by,assignee,
                     creation_date,resolved_on,priority,status,hours_spent,notes,raw_input)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', (session['user_id'], e['task_id'], e['type'], e['description'],
                      e['requested_by'], e['assignee'], e['creation_date'], e['resolved_on'],
                      e['priority'], e['status'], e['hours_spent'], e['notes'], e['raw_input']))
                added += 1
            conn.commit()
            conn.close()
        except Exception as ex:
            errors.append(f"Could not parse: {line[:60]}… ({ex})")

    if added:
        flash(f'✓ Added {added} task entr{"y" if added==1 else "ies"} successfully.', 'success')
    for err in errors:
        flash(err, 'error')

    # Always redirect to the week the input data belongs to (even if all were duplicates)
    if all_dates:
        target_fri = friday_of(max(all_dates)).isoformat()
        return redirect(f'/?week={target_fri}')
    return redirect('/')

@app.route('/delete/<int:task_id>')
@login_required
def delete_task(task_id):
    conn = db()
    conn.execute('DELETE FROM tasks WHERE id=? AND user_id=?', (task_id, session['user_id']))
    conn.commit()
    conn.close()
    flash('Entry deleted.', 'success')
    return redirect(request.referrer or '/')

@app.route('/clone/<int:task_id>')
@login_required
def clone_task(task_id):
    conn = db()
    row = conn.execute('SELECT * FROM tasks WHERE id=? AND user_id=?',
                       (task_id, session['user_id'])).fetchone()
    if row:
        conn.execute('''
            INSERT INTO tasks (user_id,task_id,type,description,requested_by,assignee,
                               creation_date,resolved_on,priority,status,hours_spent,notes,raw_input)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (session['user_id'], row['task_id'], row['type'], row['description'],
              row['requested_by'], row['assignee'], row['creation_date'], row['resolved_on'],
              row['priority'], row['status'], row['hours_spent'], row['notes'], row['raw_input']))
        conn.commit()
        flash('✓ Row cloned — original kept, copy added below.', 'success')
        # Redirect to the exact week so both original and clone are visible
        week = request.args.get('week') or friday_of(
            date.fromisoformat(row['creation_date'])).isoformat()
    else:
        flash('Row not found.', 'error')
        week = request.args.get('week', friday_of().isoformat())
    conn.close()
    return redirect(f'/?week={week}')

@app.route('/all-tasks')
@login_required
def all_tasks():
    conn = db()
    rows = conn.execute('''
        SELECT DISTINCT creation_date FROM tasks WHERE user_id=?
        ORDER BY creation_date DESC
    ''', (session['user_id'],)).fetchall()
    conn.close()

    # Build list of unique friday dates for weeks that have tasks
    seen, weeks = set(), []
    for r in rows:
        d = date.fromisoformat(r['creation_date'])
        fri = friday_of(d).isoformat()
        if fri not in seen:
            seen.add(fri)
            weeks.append(fri)

    if not weeks:
        weeks = [friday_of().isoformat()]

    sel_week = request.args.get('week', weeks[0])
    sel_fri = date.fromisoformat(sel_week)
    sel_mon, _ = week_bounds(sel_fri)

    conn = db()
    tasks = conn.execute('''
        SELECT * FROM tasks WHERE user_id=?
        AND creation_date BETWEEN ? AND ?
        ORDER BY creation_date ASC, id ASC
    ''', (session['user_id'], sel_mon.isoformat(), sel_fri.isoformat())).fetchall()
    conn.close()
    total = sum(t['hours_spent'] or 0 for t in tasks)
    return render_template_string(ALL_TASKS_T,
        tasks=tasks, weeks=weeks, sel_week=sel_week, total_hours=round(total, 1))

@app.route('/report')
@login_required
def report():
    week_str = request.args.get('week', friday_of().isoformat())
    week_fri = date.fromisoformat(week_str)
    buf = make_report(week_fri)
    fname = f"MAG_WeeklyTasks_{week_fri.isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '')
        conn = db()
        user = conn.execute('SELECT * FROM users WHERE username=?', (u,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], p):
            session.permanent   = True   # keep logged in for 30 days
            session['user_id']   = user['id']
            session['full_name'] = user['full_name']
            session['username']  = user['username']
            return redirect('/')
        flash('Invalid username or password.', 'error')
    return render_template_string(LOGIN_T)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        fn = request.form.get('full_name', '').strip()
        u  = request.form.get('username', '').strip()
        p  = request.form.get('password', '')
        if not fn or not u or not p:
            flash('All fields are required.', 'error')
            return render_template_string(REGISTER_T)
        try:
            conn = db()
            conn.execute(
                'INSERT INTO users (username, password_hash, full_name) VALUES (?,?,?)',
                (u, generate_password_hash(p, method='pbkdf2:sha256'), fn)
            )
            conn.commit()
            conn.close()
            flash('Account created! Please sign in.', 'success')
            return redirect('/login')
        except sqlite3.IntegrityError:
            flash('Username already taken. Try a different username.', 'error')
        except Exception as e:
            app.logger.error(f"Register error: {e}", exc_info=True)
            flash(f'Registration failed: {e}', 'error')
    return render_template_string(REGISTER_T)

@app.route('/clear-week/<week_fri>')
@login_required
def clear_week(week_fri):
    fri = date.fromisoformat(week_fri)
    mon, _ = week_bounds(fri)
    conn = db()
    conn.execute('''
        DELETE FROM tasks
        WHERE user_id=? AND creation_date BETWEEN ? AND ?
    ''', (session['user_id'], mon.isoformat(), fri.isoformat()))
    conn.commit()
    conn.close()
    flash(f'✓ All your entries for week {mon.strftime("%d %b")}–{fri.strftime("%d %b")} cleared.', 'success')
    return redirect(f'/?week={week_fri}')

@app.route('/update/<int:task_id>', methods=['POST'])
@login_required
def update_task(task_id):
    field = request.form.get('field')
    value = request.form.get('value', '').strip()
    allowed = {'hours_spent', 'notes', 'task_id', 'type', 'description',
               'status', 'creation_date', 'resolved_on'}
    if field not in allowed:
        return jsonify(ok=False, error='Invalid field'), 400
    if field == 'hours_spent':
        try:
            value = float(value)
            if value < 0 or value > 24:
                return jsonify(ok=False, error='Hours must be 0–24'), 400
        except ValueError:
            return jsonify(ok=False, error='Not a valid number'), 400
    elif field in ('creation_date', 'resolved_on'):
        try:
            date.fromisoformat(value)
        except ValueError:
            return jsonify(ok=False, error='Invalid date'), 400
    elif field == 'status':
        if value not in ('Completed', 'In Progress', 'Pending', 'On Hold'):
            return jsonify(ok=False, error='Invalid status'), 400
    else:
        value = value.strip() if value and value.strip() else None
    conn = db()
    conn.execute(f'UPDATE tasks SET {field}=? WHERE id=? AND user_id=?',
                 (value, task_id, session['user_id']))
    conn.commit()
    conn.close()
    return jsonify(ok=True)

@app.route('/add-row', methods=['POST'])
@login_required
def add_row():
    """Save a row entered directly in the inline table editor."""
    full_name = session['full_name']
    task_id   = request.form.get('task_id', '').strip().upper() or 'Adhoc'
    type_val  = request.form.get('type', 'General & Admin').strip() or 'General & Admin'
    desc      = request.form.get('description', '').strip()
    creation_date = request.form.get('creation_date', date.today().isoformat())
    resolved_on   = request.form.get('resolved_on', '').strip()
    priority  = request.form.get('priority', 'Low').strip()
    status    = request.form.get('status', 'Completed').strip()
    notes     = request.form.get('notes', '').strip() or None
    try:
        hours = float(request.form.get('hours_spent', 0) or 0)
    except ValueError:
        hours = 0.0
    if not resolved_on:
        cdate = date.fromisoformat(creation_date)
        resolved_on = (cdate + timedelta(days=1)).isoformat()
    conn = db()
    conn.execute('''
        INSERT INTO tasks (user_id,task_id,type,description,requested_by,assignee,
                           creation_date,resolved_on,priority,status,hours_spent,notes,raw_input)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (session['user_id'], task_id, type_val, desc, full_name, full_name,
          creation_date, resolved_on, priority, status, hours, notes, 'inline'))
    conn.commit()
    conn.close()
    week = friday_of(date.fromisoformat(creation_date)).isoformat()
    return jsonify(ok=True, week=week)

@app.route('/delete-selected', methods=['POST'])
@login_required
def delete_selected():
    ids = request.form.getlist('ids')
    if ids:
        conn = db()
        for tid in ids:
            conn.execute('DELETE FROM tasks WHERE id=? AND user_id=?', (tid, session['user_id']))
        conn.commit()
        conn.close()
        flash(f'✓ Deleted {len(ids)} entr{"y" if len(ids)==1 else "ies"}.', 'success')
    else:
        flash('No entries selected.', 'error')
    return redirect(request.referrer or '/')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("\n  MAG Weekly Tasks App running at: http://localhost:5000\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
